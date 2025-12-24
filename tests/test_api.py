"""Minimal integration tests for the FastAPI backend."""

from __future__ import annotations

from typing import Iterator

import json
import os
import sys
import io
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from fastapi.testclient import TestClient
from sqlmodel import Session, delete, select

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

os.environ.setdefault("REFDATA_DATABASE_URL", "sqlite:///:memory:")

from api.app.main import create_app
from api.app.config import Settings
from api.app.database import create_db_engine, init_db, get_session
from api.app.models import SourceSample, SystemConfig
from api.app.services.source_connections import SourceConnectionServiceError


def build_test_client() -> TestClient:
    settings = Settings(database_url="sqlite:///:memory:", match_threshold=0.55)
    app = create_app(settings)
    engine = create_db_engine(settings)
    app.state.engine = engine
    init_db(engine, settings=settings)

    def session_override() -> Iterator[Session]:
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = session_override
    return TestClient(app)


def create_sqlite_source() -> tuple[Path, tempfile.TemporaryDirectory]:
    temp_dir = tempfile.TemporaryDirectory()
    db_path = Path(temp_dir.name) / "source.db"
    connection = sqlite3.connect(db_path)
    try:
        cursor = connection.cursor()
        cursor.execute(
            "CREATE TABLE customers (id INTEGER PRIMARY KEY, name TEXT, email TEXT)"
        )
        cursor.execute(
            "CREATE VIEW customer_view AS SELECT name FROM customers"
        )
        connection.commit()
    finally:
        connection.close()

    return db_path, temp_dir


def test_health_endpoint() -> None:
    client = build_test_client()
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_match_proposal_returns_candidates() -> None:
    client = build_test_client()
    response = client.post(
        "/api/reference/propose",
        json={"raw_text": "married", "dimension": "marital_status"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["matches"]
    top_match = payload["matches"][0]
    assert top_match["canonical_label"] in {"Married", "Single"}


def test_match_proposal_falls_back_when_default_dimension_empty() -> None:
    client = build_test_client()

    config_response = client.get("/api/config")
    assert config_response.status_code == 200
    config = config_response.json()
    assert config["default_dimension"] == "general"

    response = client.post(
        "/api/reference/propose",
        json={"raw_text": "Single"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["matches"]
    assert payload["dimension"] == "marital_status"
    assert any(match["canonical_label"] == "Single" for match in payload["matches"])


def test_canonical_crud_operations() -> None:
    client = build_test_client()

    dimension_code = "test"
    dimension_response = client.post(
        "/api/reference/dimensions",
        json={
            "code": dimension_code,
            "label": "Test dimension",
            "description": "Dimension used for CRUD tests",
            "extra_fields": [
                {
                    "key": "numeric_code",
                    "label": "Numeric Code",
                    "data_type": "number",
                    "required": False,
                }
            ],
        },
    )
    assert dimension_response.status_code == 201

    create_response = client.post(
        "/api/reference/canonical",
        json={
            "dimension": dimension_code,
            "canonical_label": "Alpha",
            "description": "seed",
            "attributes": {"numeric_code": 101},
        },
    )
    assert create_response.status_code == 201
    created = create_response.json()
    canonical_id = created["id"]

    update_response = client.put(
        f"/api/reference/canonical/{canonical_id}",
        json={"canonical_label": "Alpha Prime", "attributes": {"numeric_code": 202}},
    )
    assert update_response.status_code == 200
    assert update_response.json()["canonical_label"] == "Alpha Prime"
    assert update_response.json()["attributes"]["numeric_code"] == 202

    delete_response = client.delete(f"/api/reference/canonical/{canonical_id}")
    assert delete_response.status_code == 204


def test_dimension_crud_and_extra_fields() -> None:
    client = build_test_client()

    payload = {
        "code": "region",
        "label": "Region",
        "description": "Geographic region",
        "extra_fields": [
            {
                "key": "iso_code",
                "label": "ISO Code",
                "data_type": "string",
                "required": True,
            }
        ],
    }

    create_response = client.post("/api/reference/dimensions", json=payload)
    assert create_response.status_code == 201

    list_response = client.get("/api/reference/dimensions")
    assert list_response.status_code == 200
    dimensions = list_response.json()
    assert any(dimension["code"] == "region" for dimension in dimensions)

    update_response = client.put(
        "/api/reference/dimensions/region",
        json={
            "label": "Region Updated",
            "extra_fields": [
                {
                    "key": "iso_code",
                    "label": "ISO",
                    "data_type": "string",
                    "required": True,
                },
                {
                    "key": "area_sq_km",
                    "label": "Area (km²)",
                    "data_type": "number",
                    "required": False,
                },
            ],
        },
    )
    assert update_response.status_code == 200
    updated = update_response.json()
    assert updated["label"] == "Region Updated"
    assert len(updated["extra_fields"]) == 2

    delete_response = client.delete("/api/reference/dimensions/region")
    assert delete_response.status_code == 204

    confirm_response = client.get("/api/reference/dimensions")
    assert confirm_response.status_code == 200
    assert all(dimension["code"] != "region" for dimension in confirm_response.json())


def test_dimension_relation_flow() -> None:
    client = build_test_client()

    for code, label in (("region", "Region"), ("district", "District")):
        client.post(
            "/api/reference/dimensions",
            json={
                "code": code,
                "label": label,
                "description": f"{label} dimension",
                "extra_fields": [],
            },
        )

    region = client.post(
        "/api/reference/canonical",
        json={"dimension": "region", "canonical_label": "North"},
    ).json()
    district = client.post(
        "/api/reference/canonical",
        json={"dimension": "district", "canonical_label": "North-01"},
    ).json()

    relation_response = client.post(
        "/api/reference/dimension-relations",
        json={
            "label": "Region to District",
            "parent_dimension_code": "region",
            "child_dimension_code": "district",
            "description": "Regions contain multiple districts",
        },
    )
    assert relation_response.status_code == 201
    relation_id = relation_response.json()["id"]

    link_response = client.post(
        f"/api/reference/dimension-relations/{relation_id}/links",
        json={
            "parent_canonical_id": region["id"],
            "child_canonical_id": district["id"],
        },
    )
    assert link_response.status_code == 201

    links = client.get(
        f"/api/reference/dimension-relations/{relation_id}/links"
    ).json()
    assert any(link["parent_canonical_id"] == region["id"] for link in links)

    delete_link = client.delete(
        f"/api/reference/dimension-relations/{relation_id}/links/{links[0]['id']}"
    )
    assert delete_link.status_code == 204

    delete_relation = client.delete(f"/api/reference/dimension-relations/{relation_id}")
    assert delete_relation.status_code == 204


def test_bulk_import_csv_and_excel() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "bulk",
            "label": "Bulk Dimension",
            "description": "Dimension used for bulk import",
            "extra_fields": [
                {
                    "key": "code",
                    "label": "Code",
                    "data_type": "string",
                    "required": False,
                }
            ],
        },
    )

    csv_content = "dimension,label,code\nbulk,Value A,A1\n,Value B,A2\n"
    csv_response = client.post(
        "/api/reference/canonical/import",
        data={"dimension": "bulk"},
        files={"file": ("values.csv", csv_content.encode("utf-8"), "text/csv")},
    )
    assert csv_response.status_code == 201
    payload = csv_response.json()
    assert len(payload["created"]) == 2
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    assert not payload["errors"]

    dataframe = pd.DataFrame(
        [
            {"label": "Value C", "description": "Excel row", "code": "A3"},
            {"label": "Value D", "description": "Excel row", "code": "A4"},
        ]
    )
    excel_buffer = io.BytesIO()
    dataframe.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    excel_response = client.post(
        "/api/reference/canonical/import",
        data={"dimension": "bulk"},
        files={
            "file": (
                "values.xlsx",
                excel_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert excel_response.status_code == 201
    excel_payload = excel_response.json()
    assert len(excel_payload["created"]) == 2
    assert excel_payload["updated"] == []
    assert excel_payload["duplicates"] == []
    assert not excel_payload["errors"]


def test_bulk_import_excel_with_metadata_and_multiple_sheets() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "bulk_multi",
            "label": "Bulk multi-sheet dimension",
            "description": "Dimension used for Excel metadata parsing",
            "extra_fields": [
                {
                    "key": "code",
                    "label": "Code",
                    "data_type": "string",
                    "required": False,
                }
            ],
        },
    )

    workbook = Workbook()
    metadata = workbook.active
    metadata.title = "Metadata"
    metadata["A1"] = "Dataset"
    metadata["B1"] = "Canonical values"

    sheet = workbook.create_sheet("2024 Data")
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Reference data extract"
    sheet["A3"] = "Dimension"
    sheet["B3"] = "Canonical Label"
    sheet["C3"] = "Code"
    sheet["A4"] = "bulk_multi"
    sheet["B4"] = "Excel value one"
    sheet["C4"] = "EX-1"
    sheet["A5"] = "bulk_multi"
    sheet["B5"] = "Excel value two"
    sheet["C5"] = "EX-2"

    notes = workbook.create_sheet("Notes")
    notes["A1"] = "Generated"
    notes["B1"] = "2024"

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    response = client.post(
        "/api/reference/canonical/import",
        data={"dimension": "bulk_multi"},
        files={
            "file": (
                "bulk_multi.xlsx",
                excel_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["errors"] == []
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    labels = [entry["canonical_label"] for entry in payload["created"]]
    assert {"Excel value one", "Excel value two"} == set(labels)


def test_bulk_import_preview_allows_sheet_selection() -> None:
    client = build_test_client()

    workbook = Workbook()
    cities = workbook.active
    cities.title = "Cities"
    cities.append(["Dimension", "Canonical Label"])
    cities.append(["region", "City One"])

    regions = workbook.create_sheet("Regions")
    regions.append(["Dimension", "Canonical Label"])
    regions.append(["region", "Region One"])

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    preview_response = client.post(
        "/api/reference/canonical/import/preview",
        files={
            "file": (
                "multi.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert set(preview_payload["available_sheets"]) == {"Cities", "Regions"}
    assert preview_payload["selected_sheet"] in {"Cities", "Regions"}

    preview_regions = client.post(
        "/api/reference/canonical/import/preview",
        data={"sheet": "Regions"},
        files={
            "file": (
                "multi.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview_regions.status_code == 200
    preview_regions_payload = preview_regions.json()
    assert preview_regions_payload["selected_sheet"] == "Regions"
    assert any(
        "Region One" in column.get("sample", [])
        for column in preview_regions_payload["columns"]
    )


def test_bulk_import_csv_with_preface_metadata() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "bulk_preface",
            "label": "Bulk metadata dimension",
            "description": "Dimension used for CSV metadata parsing",
            "extra_fields": [
                {
                    "key": "code",
                    "label": "Code",
                    "data_type": "string",
                    "required": False,
                }
            ],
        },
    )

    csv_content = (
        "Report generated,2024-05-11,\n"
        "Owner,Data Steward,\n"
        "Canonical Label,Description,Code\n"
        "CSV value one,Imported from CSV,CS-1\n"
        "CSV value two,Imported from CSV,CS-2\n"
    )

    response = client.post(
        "/api/reference/canonical/import",
        data={"dimension": "bulk_preface"},
        files={
            "file": ("bulk_preface.csv", csv_content.encode("utf-8"), "text/csv")
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["errors"] == []
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    created = {entry["canonical_label"] for entry in payload["created"]}
    assert created == {"CSV value one", "CSV value two"}


def test_bulk_import_accepts_canonical_value_header() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "bulk_headers",
            "label": "Bulk header dimension",
            "description": "Dimension used to verify header detection",
            "extra_fields": [
                {
                    "key": "numeric_code",
                    "label": "Numeric Code",
                    "data_type": "number",
                    "required": False,
                }
            ],
        },
    )

    csv_content = (
        "dimension name,canonical value,long description,Numeric Code\n"
        "bulk_headers,Header Label,Additional context,7\n"
    )

    response = client.post(
        "/api/reference/canonical/import",
        data={"dimension": "bulk_headers"},
        files={
            "file": (
                "headers.csv",
                csv_content.encode("utf-8"),
                "text/csv",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["errors"] == []
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    assert len(payload["created"]) == 1
    created_entry = payload["created"][0]
    assert created_entry["canonical_label"] == "Header Label"
    assert created_entry["description"] == "Additional context"
    assert created_entry["attributes"]["numeric_code"] == 7


def test_bulk_import_preview_suggests_mapping() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "region",
            "label": "Region",
            "extra_fields": [
                {
                    "key": "numeric_code",
                    "label": "Numeric Code",
                    "data_type": "number",
                }
            ],
        },
    )

    csv_content = "Region Label,Numeric Code\nAbu Dhabi,01\nDubai,02\n"
    response = client.post(
        "/api/reference/canonical/import/preview",
        files={"file": ("regions.csv", csv_content.encode("utf-8"), "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["columns"][0]["suggested_role"] in {"label", "dimension"}
    assert any(
        column.get("suggested_attribute_key") == "numeric_code"
        for column in payload["columns"]
    )
    assert payload["proposed_dimension"] is not None


def test_bulk_import_accepts_explicit_mapping() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "region",
            "label": "Region",
            "extra_fields": [
                {
                    "key": "numeric_code",
                    "label": "Numeric Code",
                    "data_type": "number",
                }
            ],
        },
    )

    csv_content = "Region Name,Code\nAbu Dhabi,01\nDubai,02\n"
    mapping_payload = {
        "label": "Region Name",
        "default_dimension": "region",
        "attributes": {"numeric_code": "Code"},
    }

    response = client.post(
        "/api/reference/canonical/import",
        data={"mapping": json.dumps(mapping_payload)},
        files={"file": ("regions.csv", csv_content.encode("utf-8"), "text/csv")},
    )

    assert response.status_code == 201
    payload = response.json()
    assert len(payload["created"]) == 2
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    labels = {item["canonical_label"] for item in payload["created"]}
    assert labels == {"Abu Dhabi", "Dubai"}


def test_bulk_import_creates_dimension_from_mapping_definition() -> None:
    client = build_test_client()

    csv_content = "City,Code\nSharjah,03\n"
    mapping_payload = {
        "label": "City",
        "default_dimension": "city",
        "attributes": {"numeric_code": "Code"},
        "dimension_definition": {
            "code": "city",
            "label": "City",
            "extra_fields": [
                {
                    "key": "numeric_code",
                    "label": "Numeric Code",
                    "data_type": "string",
                }
            ],
        },
    }

    response = client.post(
        "/api/reference/canonical/import",
        data={"mapping": json.dumps(mapping_payload)},
        files={"file": ("cities.csv", csv_content.encode("utf-8"), "text/csv")},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["errors"] == []
    assert payload["updated"] == []
    assert payload["duplicates"] == []
    assert payload["created"][0]["dimension"] == "city"

    dimension_response = client.get("/api/reference/dimensions")
    assert dimension_response.status_code == 200
    dimension_payloads = dimension_response.json()
    city_dimension = next(item for item in dimension_payloads if item["code"] == "city")
    assert any(field["key"] == "numeric_code" for field in city_dimension["extra_fields"])


def test_bulk_import_duplicate_resolution_strategies() -> None:
    client = build_test_client()

    client.post(
        "/api/reference/dimensions",
        json={
            "code": "region",
            "label": "Region",
            "description": "Region dimension",
            "extra_fields": [],
        },
    )

    client.post(
        "/api/reference/canonical",
        json={"dimension": "region", "canonical_label": "Abu Dhabi", "description": "Original"},
    )

    csv_content = "dimension,label,description\nregion,Abu Dhabi,Revised\nregion,Dubai,New\n"

    dry_run = client.post(
        "/api/reference/canonical/import",
        data={"dry_run": "true"},
        files={"file": ("duplicates.csv", csv_content.encode("utf-8"), "text/csv")},
    )
    assert dry_run.status_code == 201
    dry_payload = dry_run.json()
    assert len(dry_payload["duplicates"]) == 1
    assert dry_payload["created"] == []
    assert dry_payload["updated"] == []

    pending = client.post(
        "/api/reference/canonical/import",
        files={"file": ("duplicates.csv", csv_content.encode("utf-8"), "text/csv")},
    )
    assert pending.status_code == 201
    pending_payload = pending.json()
    assert len(pending_payload["duplicates"]) == 1
    assert pending_payload["created"] == []
    assert pending_payload["updated"] == []

    skip = client.post(
        "/api/reference/canonical/import",
        data={"duplicate_strategy": "skip"},
        files={"file": ("duplicates.csv", csv_content.encode("utf-8"), "text/csv")},
    )
    assert skip.status_code == 201
    skip_payload = skip.json()
    assert len(skip_payload["created"]) == 1
    assert skip_payload["created"][0]["canonical_label"] == "Dubai"
    assert skip_payload["updated"] == []
    assert skip_payload["duplicates"] == []
    assert any("Skipped" in message for message in skip_payload["errors"])

    canonical_list = client.get("/api/reference/canonical").json()
    abu_dhabi = next(item for item in canonical_list if item["canonical_label"] == "Abu Dhabi")
    assert abu_dhabi["description"] == "Original"

    update_csv = "dimension,label,description\nregion,Abu Dhabi,Updated\n"
    update = client.post(
        "/api/reference/canonical/import",
        data={"duplicate_strategy": "update"},
        files={"file": ("update.csv", update_csv.encode("utf-8"), "text/csv")},
    )
    assert update.status_code == 201
    update_payload = update.json()
    assert update_payload["created"] == []
    assert len(update_payload["updated"]) == 1
    assert update_payload["updated"][0]["description"] == "Updated"
    assert update_payload["errors"] == []

    final_list = client.get("/api/reference/canonical").json()
    updated_entry = next(item for item in final_list if item["canonical_label"] == "Abu Dhabi")
    assert updated_entry["description"] == "Updated"


def test_source_mapping_flow() -> None:
    client = build_test_client()

    # Create a source connection
    connection_response = client.post(
        "/api/source/connections",
        json={
            "name": "crm",
            "db_type": "postgres",
            "host": "localhost",
            "port": 5432,
            "database": "crm",
            "username": "svc",
        },
    )
    assert connection_response.status_code == 201
    connection_id = connection_response.json()["id"]

    # Map a field to the marital_status dimension
    mapping_response = client.post(
        f"/api/source/connections/{connection_id}/mappings",
        json={
            "source_table": "customers",
            "source_field": "marital",
            "ref_dimension": "marital_status",
        },
    )
    assert mapping_response.status_code == 201
    mapping_id = mapping_response.json()["id"]

    # Ingest raw values (one typo, one exact match)
    ingest_response = client.post(
        f"/api/source/connections/{connection_id}/samples",
        json={
            "source_table": "customers",
            "source_field": "marital",
            "values": [
                {"raw_value": "Singel", "occurrence_count": 3, "dimension": "marital_status"},
                {"raw_value": "Married", "occurrence_count": 5, "dimension": "marital_status"},
            ],
        },
    )
    assert ingest_response.status_code == 201

    # Compute match statistics
    stats_response = client.get(f"/api/source/connections/{connection_id}/match-stats")
    assert stats_response.status_code == 200
    stats = stats_response.json()
    assert stats
    stat = next(item for item in stats if item["mapping_id"] == mapping_id)
    assert stat["unmatched_values"] >= 1
    assert stat["total_values"] == 8
    assert stat["top_matched"]
    assert any(match["raw_value"] == "Married" for match in stat["top_matched"])

    # Retrieve unmatched values (should include "Singel")
    unmatched_response = client.get(f"/api/source/connections/{connection_id}/unmatched")
    assert unmatched_response.status_code == 200
    unmatched = unmatched_response.json()
    assert any(record["raw_value"] == "Singel" for record in unmatched)

    # Approve a mapping for the typo using the existing canonical "Single"
    canonical_response = client.get("/api/reference/canonical")
    canonical_options = canonical_response.json()
    single = next(item for item in canonical_options if item["canonical_label"] == "Single")

    create_mapping_response = client.post(
        f"/api/source/connections/{connection_id}/value-mappings",
        json={
            "source_table": "customers",
            "source_field": "marital",
            "raw_value": "Singel",
            "canonical_id": single["id"],
            "status": "approved",
        },
    )
    assert create_mapping_response.status_code == 201

    refreshed_stats = client.get(
        f"/api/source/connections/{connection_id}/match-stats"
    ).json()
    updated_stat = next(item for item in refreshed_stats if item["mapping_id"] == mapping_id)
    assert any(match["raw_value"] == "Singel" for match in updated_stat["top_matched"])

    # The value should disappear from unmatched results
    unmatched_after = client.get(f"/api/source/connections/{connection_id}/unmatched").json()
    assert not any(record["raw_value"] == "Singel" for record in unmatched_after)

    # Connection specific mapping list
    connection_mappings = client.get(
        f"/api/source/connections/{connection_id}/value-mappings"
    )
    assert connection_mappings.status_code == 200
    assert connection_mappings.json()

    # Global mapping list
    all_mappings = client.get("/api/source/value-mappings")
    assert all_mappings.status_code == 200
    assert all_mappings.json()


def test_source_sample_endpoint_returns_distinct_values() -> None:
    client = build_test_client()

    connection_response = client.post(
        "/api/source/connections",
        json={
            "name": "analytics",
            "db_type": "postgres",
            "host": "localhost",
            "port": 5432,
            "database": "analytics",
            "username": "svc",
        },
    )
    assert connection_response.status_code == 201
    connection_id = connection_response.json()["id"]

    engine = client.app.state.engine
    assert engine is not None

    with Session(engine) as session:
        session.add_all(
            [
                SourceSample(
                    source_connection_id=connection_id,
                    source_table="customers",
                    source_field="email",
                    dimension=None,
                    raw_value="alice@example.com",
                    occurrence_count=4,
                    last_seen_at=datetime(2024, 1, 5, 12, 0, tzinfo=timezone.utc),
                ),
                SourceSample(
                    source_connection_id=connection_id,
                    source_table="customers",
                    source_field="email",
                    dimension="contact",
                    raw_value="alice@example.com",
                    occurrence_count=3,
                    last_seen_at=datetime(2024, 2, 15, 8, 0, tzinfo=timezone.utc),
                ),
                SourceSample(
                    source_connection_id=connection_id,
                    source_table="customers",
                    source_field="email",
                    dimension=None,
                    raw_value="bob@example.com",
                    occurrence_count=2,
                    last_seen_at=datetime(2024, 1, 20, 10, 30, tzinfo=timezone.utc),
                ),
            ]
        )
        session.commit()

    response = client.get(
        f"/api/source/connections/{connection_id}/samples",
        params={"source_table": "customers", "source_field": "email"},
    )
    assert response.status_code == 200
    payload = response.json()

    assert len(payload) == 2
    alice = next(item for item in payload if item["raw_value"] == "alice@example.com")
    bob = next(item for item in payload if item["raw_value"] == "bob@example.com")

    assert alice["occurrence_count"] == 7
    assert alice["dimension"] == "contact"
    assert alice["last_seen_at"].startswith("2024-02-15T08:00:00")

    assert bob["occurrence_count"] == 2
    assert bob["dimension"] is None


def test_capture_mapping_samples_ingests_source_values() -> None:
    client = build_test_client()
    db_path, temp_dir = create_sqlite_source()
    try:
        connection = sqlite3.connect(db_path)
        try:
            cursor = connection.cursor()
            cursor.executemany(
                "INSERT INTO customers (name, email) VALUES (?, ?)",
                [
                    ("Alice", "alice@example.com"),
                    ("Bob", "bob@example.com"),
                    ("Bob", "bob@example.com"),
                ],
            )
            connection.commit()
        finally:
            connection.close()

        connection_response = client.post(
            "/api/source/connections",
            json={
                "name": "sqlite-warehouse",
                "db_type": "sqlite",
                "host": "localhost",
                "port": 5432,
                "database": str(db_path),
                "username": "ignored",
            },
        )
        assert connection_response.status_code == 201
        connection_id = connection_response.json()["id"]

        mapping_response = client.post(
            f"/api/source/connections/{connection_id}/mappings",
            json={
                "source_table": "customers",
                "source_field": "name",
                "ref_dimension": "customer_name",
            },
        )
        assert mapping_response.status_code == 201
        mapping_id = mapping_response.json()["id"]

        capture_response = client.post(
            f"/api/source/connections/{connection_id}/mappings/{mapping_id}/capture",
        )
        assert capture_response.status_code == 201
        payload = capture_response.json()
        assert payload
        bob_sample = next(item for item in payload if item["raw_value"] == "Bob")
        assert bob_sample["occurrence_count"] == 2

        stats_response = client.get(f"/api/source/connections/{connection_id}/match-stats")
        assert stats_response.status_code == 200
        stats = stats_response.json()
        assert stats[0]["total_values"] == 3
    finally:
        temp_dir.cleanup()
def test_source_connection_test_endpoint_success() -> None:
    client = build_test_client()
    db_path, temp_dir = create_sqlite_source()
    try:
        response = client.post(
            "/api/source/connections/test",
            json={
                "name": "analytics-sqlite",
                "db_type": "sqlite",
                "host": "localhost",
                "port": 5432,
                "database": str(db_path),
                "username": "ignored",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["success"] is True
        assert "Connection succeeded" in payload["message"]
        assert isinstance(payload["latency_ms"], float)

        failure = client.post(
            "/api/source/connections/test",
            json={
                "db_type": "sqlite",
                "host": "localhost",
                "port": 5432,
                "database": str(db_path),
                "username": "ignored",
                "options": '{"schema":',
            },
        )
        assert failure.status_code == 400
        assert "detail" in failure.json()
    finally:
        temp_dir.cleanup()


def test_source_connection_metadata_and_existing_test() -> None:
    client = build_test_client()
    db_path, temp_dir = create_sqlite_source()
    try:
        create_response = client.post(
            "/api/source/connections",
            json={
                "name": "warehouse-sqlite",
                "db_type": "sqlite",
                "host": "localhost",
                "port": 5432,
                "database": str(db_path),
                "username": "ignored",
            },
        )
        assert create_response.status_code == 201
        connection_id = create_response.json()["id"]

        detail_response = client.get(f"/api/source/connections/{connection_id}")
        assert detail_response.status_code == 200
        detail = detail_response.json()
        assert detail["name"] == "warehouse-sqlite"
        assert detail["database"].endswith("source.db")

        tables_response = client.get(f"/api/source/connections/{connection_id}/tables")
        assert tables_response.status_code == 200
        tables = tables_response.json()
        assert any(table["name"] == "customers" for table in tables)
        assert any(table["type"] == "view" for table in tables)

        fields_response = client.get(
            f"/api/source/connections/{connection_id}/tables/customers/fields",
        )
        assert fields_response.status_code == 200
        fields = fields_response.json()
        assert any(field["name"] == "email" for field in fields)

        schema_fields_response = client.get(
            f"/api/source/connections/{connection_id}/tables/customer_view/fields?schema=main",
        )
        assert schema_fields_response.status_code == 200
        schema_fields = schema_fields_response.json()
        assert any(field["name"] == "name" for field in schema_fields)

        success = client.post(f"/api/source/connections/{connection_id}/test")
        assert success.status_code == 200

        override_failure = client.post(
            f"/api/source/connections/{connection_id}/test",
            json={"db_type": "oracle"},
        )
        assert override_failure.status_code == 400

    finally:
        temp_dir.cleanup()


def test_source_tables_surface_connection_errors(monkeypatch: pytest.MonkeyPatch) -> None:
    client = build_test_client()
    db_path, temp_dir = create_sqlite_source()
    try:
        create_response = client.post(
            "/api/source/connections",
            json={
                "name": "unreachable-db",
                "db_type": "sqlite",
                "host": "localhost",
                "port": 5432,
                "database": str(db_path),
                "username": "ignored",
            },
        )
        assert create_response.status_code == 201
        connection_id = create_response.json()["id"]

        def _fail(*_: object, **__: object) -> None:
            raise SourceConnectionServiceError("connection is bad: Name or service not known")

        monkeypatch.setattr("api.app.routes.source.service_list_tables", _fail)

        response = client.get(f"/api/source/connections/{connection_id}/tables")
        assert response.status_code == 400
        assert "connection is bad" in response.json()["detail"]
    finally:
        temp_dir.cleanup()


def test_config_endpoint_auto_seeds_when_missing(tmp_path) -> None:
    db_path = tmp_path / "config.db"
    settings = Settings(database_url=f"sqlite:///{db_path}")
    app = create_app(settings)
    engine = create_db_engine(settings)
    app.state.engine = engine
    init_db(engine, settings=settings)

    def session_override() -> Iterator[Session]:
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = session_override

    with TestClient(app) as client:
        with Session(engine) as session:
            session.exec(delete(SystemConfig))
            session.commit()

        response = client.get("/api/config")
        assert response.status_code == 200
        payload = response.json()
        assert payload["default_dimension"] == settings.default_dimension
        assert payload["match_threshold"] == settings.match_threshold
        assert payload["llm_mode"] == settings.llm_mode

        with Session(engine) as session:
            config = session.exec(select(SystemConfig)).first()
            assert config is not None
            assert config.default_dimension == settings.default_dimension
            assert config.llm_mode == settings.llm_mode


def test_seeded_source_connection_available() -> None:
    client = build_test_client()
    response = client.get("/api/source/connections")
    assert response.status_code == 200
    payload = response.json()
    assert any(connection["name"] == "Target Demo Warehouse" for connection in payload)


def test_match_stats_without_samples_returns_empty_totals() -> None:
    client = build_test_client()

    connection_response = client.post(
        "/api/source/connections",
        json={
            "name": "Demo Warehouse",
            "db_type": "postgres",
            "host": "localhost",
            "port": 5432,
            "database": "targetdb",
            "username": "svc",
            "password": None,
            "options": None,
        },
    )
    assert connection_response.status_code == 201
    connection_id = connection_response.json()["id"]

    mapping_response = client.post(
        f"/api/source/connections/{connection_id}/mappings",
        json={
            "source_table": "public.customers",
            "source_field": "marital_status",
            "ref_dimension": "marital_status",
            "description": "Demo mapping",
        },
    )
    assert mapping_response.status_code == 201

    response = client.get(f"/api/source/connections/{connection_id}/match-stats")
    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    stats = payload[0]
    assert stats["total_values"] == 0
    assert stats["matched_values"] == 0
    assert stats["unmatched_values"] == 0
    assert stats["match_rate"] == 0.0
    assert stats["top_unmatched"] == []
